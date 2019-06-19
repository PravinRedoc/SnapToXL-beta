import win32clipboard
import io
import codecs
import PIL
from PIL import ImageGrab
import xlsxwriter
import openpyxl
from openpyxl import load_workbook
from xlwt import Workbook 
import time
addoff = 1

def createExcel():
    filename=e.get()
    filepath=str(z.get()).replace("\\","\\\\")
    # zz=str(ss).replace("\\","\\\\")
    workbook=xlsxwriter.Workbook(filepath+'\\'+filename+'.xlsx')
    # add_sheet is used to create sheet. 
    workbook.add_worksheet()
    workbook.close()

def saveToexcel(img,imgname,offset):
    
    
    filename=e.get()
    filepath=str(z.get()).replace("\\","\\\\")
    wb=load_workbook(filepath+'\\'+filename+'.xlsx')
    # select demo.xlsx
    ws = wb.worksheets[0]
    img.save(imgname,'PNG') 
    img = openpyxl.drawing.image.Image(imgname)
    addoff=offset
    img.anchor = 'A'+str(addoff)
    ws.add_image(img)
    wb.save(filepath+'\\'+filename+'.xlsx')
    return addoff+42

def grabScreenShot():
    global addoff
    root.wm_state('iconic')
    time.sleep(2)
    filepath=str(z.get()).replace("\\","\\\\")
    im = ImageGrab.grab() 
    addoff=saveToexcel(im,filepath+'\\lastcopiedimg.png',addoff)
    #root.state('zoomed')
    #time.sleep(2)
    root.wm_state('normal')
        


from tkinter import *

root = Tk()
root.title("SnapToXL beta")
root.geometry("210x100")
root.resizable(False,False)
Label(root,text = 'Enter File Path').grid(row=0)
z = Entry(root)
z.grid(row=0,column=1)
z.focus_set()


Label(root,text = 'Enter File Name').grid(row=1)
e = Entry(root)
e.grid(row=1,column=1)
e.focus_set()

button = Button(root, text='Create Excel', command = createExcel)
button.grid(row=3)

Button(root,text='Copy to Excel',command = grabScreenShot).grid(row=3,column=1)

#Button(root,text='Copy to Excel').grid(row=3,column=2)


root.mainloop()