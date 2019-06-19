# -*- coding: utf-8 -*-
"""
Created on Tue Jun 18 15:48:22 2019

@author: praveen-ch
"""

# import load_workbook
#from openpyxl import load_workbook
## set file path
#filepath='C:\\Automation\\images1.xlsx'
## load demo.xlsx 
#wb=load_workbook(filepath)
## select demo.xlsx
#sheet=wb.active
## set value for cell A1=1
#sheet['A2'] = "A for APPPLE"
## set value for cell B2=2
#sheet.cell(row=2, column=2).value = 2
## save workbook 
#wb.save(filepath)
import xlwt 
from xlwt import Workbook 

import openpyxl
from openpyxl import load_workbook
# set file path
filepath='C:\\Automation\\images.xlsx'
# load demo.xlsx 
wb=load_workbook(filepath)
# select demo.xlsx
ws = wb.worksheets[0]
img = openpyxl.drawing.image.Image('C:\\Automation\\somefile4.png' )
img.anchor = 'B42'
ws.add_image(img)
wb.save(filepath)



#import openpyxl
#wb = openpyxl.Workbook()
#ws = wb.worksheets[0]
#img = openpyxl.drawing.image.Image('test.jpg')
#img.anchor = 'A1'
#ws.add_image(img)
#wb.save('out.xlsx')
