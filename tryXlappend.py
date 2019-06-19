# -*- coding: utf-8 -*-


"""
Created on Tue Jun 18 15:17:30 2019

@author: praveen-ch
"""

import pandas
from openpyxl import load_workbook

book = load_workbook('C:\\Automation\\images1.xlsx')
writer = pandas.ExcelWriter('Masterfile.xlsx', engine='openpyxl') 
writer.book = book
writer.sheets = dict((ws.title, ws) for ws in book.worksheets)

data_filtered.to_excel(writer, "Main", cols=['Diff1', 'Diff2'])

writer.save()


filepath='C:\\Automation\\images.xlsx'
# load demo.xlsx 
wb=load_workbook(filepath)
# select demo.xlsx
ws = wb.worksheets[0]
img = openpyxl.drawing.image.Image('C:\\Automation\\somefile4.png' )
img.anchor = 'B42'
ws.add_image(img)
wb.save(filepath)
